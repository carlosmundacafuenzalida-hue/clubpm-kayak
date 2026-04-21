"""
Script de importación única desde planilla Excel a Supabase.

Estructura esperada (archivo: scripts/planilla.xlsx):
  Hoja "Socios":   RUT | Nombre | Telefono | FechaIngreso
  Hoja "Pagos":    RUT | Mes (YYYY-MM) | Monto | Glosa | Fecha
  Hoja "Gastos":   Fecha | Monto | Glosa
  Hoja "Ingresos": Fecha | Monto | Glosa

Uso:
  cd scripts && pip install -r requirements.txt
  python import_planilla.py
"""
import os, sys
from datetime import datetime
from dotenv import load_dotenv
from supabase import create_client
import openpyxl

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env.local'))

SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.environ.get('NEXT_PUBLIC_SUPABASE_ANON_KEY')
ADMIN_RUT    = os.environ.get('NEXT_PUBLIC_ADMIN_RUT', 'admin')
ARCHIVO      = os.path.join(os.path.dirname(__file__), 'planilla.xlsx')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Faltan NEXT_PUBLIC_SUPABASE_URL o NEXT_PUBLIC_SUPABASE_ANON_KEY en .env.local")
    sys.exit(1)

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


def normalizar_rut(rut) -> str:
    clean = str(rut).strip().replace('.', '').replace('-', '').lower()
    if len(clean) < 2:
        return clean
    return f"{clean[:-1]}-{clean[-1]}"


def parse_fecha(valor, fallback: str = None) -> str:
    if valor is None:
        return fallback or datetime.today().strftime('%Y-%m-%d')
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d')
    s = str(valor).strip()[:10]
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return fallback or datetime.today().strftime('%Y-%m-%d')


def importar_socios(ws):
    print("\n=== Socios ===")
    ok = skip = err = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:
            continue
        rut, nombre, telefono, fecha_ingreso = (list(row) + [None]*4)[:4]
        rut_norm = normalizar_rut(rut)
        nombre = str(nombre).strip() if nombre else ''
        if not nombre:
            print(f"  SKIP fila {i}: nombre vacío"); skip += 1; continue
        if supabase.table('socios').select('id').eq('rut', rut_norm).execute().data:
            print(f"  SKIP {rut_norm} (ya existe)"); skip += 1; continue
        try:
            supabase.table('socios').insert({
                'rut': rut_norm, 'nombre': nombre,
                'telefono': str(telefono).strip() if telefono else None,
                'fecha_ingreso': parse_fecha(fecha_ingreso, '2024-08-01'),
                'activo': True, 'es_admin': False,
            }).execute()
            print(f"  OK  {rut_norm} - {nombre}"); ok += 1
        except Exception as e:
            print(f"  ERR {rut_norm}: {e}"); err += 1
    print(f"  → {ok} insertados, {skip} saltados, {err} errores")


def importar_pagos(ws):
    print("\n=== Pagos ===")
    socios_map = {s['rut']: s['id'] for s in supabase.table('socios').select('id, rut').execute().data}
    ok = skip = err = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0] or not row[2]:
            continue
        rut, mes, monto, glosa, fecha = (list(row) + [None]*5)[:5]
        rut_norm = normalizar_rut(rut)
        socio_id = socios_map.get(rut_norm)
        if not socio_id:
            print(f"  WARN fila {i}: {rut_norm} no encontrado"); skip += 1; continue
        if isinstance(mes, datetime):
            mes_str = mes.strftime('%Y-%m') + '-01'
        else:
            s = str(mes).strip()[:7]
            mes_str = s + '-01' if len(s) == 7 else None
        try:
            supabase.table('movimientos').insert({
                'tipo': 'pago_cuota',
                'fecha_registro': parse_fecha(fecha, mes_str),
                'socio_id': socio_id, 'mes_cuota': mes_str,
                'monto': float(monto),
                'glosa': str(glosa).strip() if glosa else 'Pago cuota',
                'creado_por': ADMIN_RUT,
            }).execute()
            print(f"  OK  {rut_norm} | {mes_str} | ${monto}"); ok += 1
        except Exception as e:
            print(f"  ERR fila {i}: {e}"); err += 1
    print(f"  → {ok} insertados, {skip} saltados, {err} errores")


def importar_movimientos(ws, tipo: str):
    label = 'gastos' if tipo == 'gasto' else 'ingresos extra'
    print(f"\n=== {label.capitalize()} ===")
    ok = skip = err = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[1]:
            continue
        fecha, monto, glosa = (list(row) + [None]*3)[:3]
        try:
            supabase.table('movimientos').insert({
                'tipo': tipo,
                'fecha_registro': parse_fecha(fecha),
                'monto': float(monto),
                'glosa': str(glosa).strip() if glosa else tipo,
                'creado_por': ADMIN_RUT,
            }).execute()
            print(f"  OK  {parse_fecha(fecha)} | ${monto} | {glosa}"); ok += 1
        except Exception as e:
            print(f"  ERR fila {i}: {e}"); err += 1
    print(f"  → {ok} insertados, {skip} saltados, {err} errores")


if __name__ == '__main__':
    if not os.path.exists(ARCHIVO):
        print(f"ERROR: No se encontró {ARCHIVO}")
        print("Copia tu planilla como 'scripts/planilla.xlsx' y vuelve a ejecutar.")
        sys.exit(1)
    wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
    print(f"Hojas encontradas: {wb.sheetnames}")
    if 'Socios'   in wb.sheetnames: importar_socios(wb['Socios'])
    if 'Pagos'    in wb.sheetnames: importar_pagos(wb['Pagos'])
    if 'Gastos'   in wb.sheetnames: importar_movimientos(wb['Gastos'], 'gasto')
    if 'Ingresos' in wb.sheetnames: importar_movimientos(wb['Ingresos'], 'ingreso_extra')
    print("\n✓ Importación completada.")
